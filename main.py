import schedule
import time
import openpyxl
from openpyxl import Workbook

def create_excel_sheet():
    wb = Workbook()
    sheet = wb.active
#dkjfgg
    sheet["A1"] = "Date"
    sheet["B1"] = "Tasks Completed"
    sheet["C1"] = "Money Spent ($)"
    sheet["D1"] = "Study Duration (hours)"

    # Save the workbook with a name that includes the current date
    filename = f"Daily_Report_{time.strftime('%Y-%m-%d')}.xlsx"
    wb.save(filename)

def daily_task_report():
    tasks = input("Tasks Completed: ")
    money_spent = input("Money Spent (in dollars): ")
    study_duration = input("Study Duration (in hours): ")

    # Load the existing workbook
    filename = f"Daily_Report_{time.strftime('%Y-%m-%d')}.xlsx"
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active

    # Find the first empty row in the sheet
    row = sheet.max_row + 1

    # Add the daily tasks to the sheet
    sheet.cell(row=row, column=1, value=time.strftime('%Y-%m-%d'))
    sheet.cell(row=row, column=2, value=tasks)
    sheet.cell(row=row, column=3, value=money_spent)
    sheet.cell(row=row, column=4, value=study_duration)

    # Save the changes
    wb.save(filename)

# Schedule the create_excel_sheet function to be executed at 10 PM every day
schedule.every().day.at("22:00").do(create_excel_sheet)

while True:
    # Check if there are any pending jobs from the schedule and run them
    schedule.run_pending()
    time.sleep(1)
print("Done")